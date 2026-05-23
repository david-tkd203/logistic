import json
from decimal import Decimal
from datetime import date, datetime, timedelta
import time

from django.db import transaction
from django.http import StreamingHttpResponse
from django.views.decorators.http import require_GET
from rest_framework import viewsets, status, mixins
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from django.db.models import Sum, Count, Q, F, FloatField, ExpressionWrapper

from .models import (
    Sucursal, Categoria, Producto, Proveedor, ProductoProveedor,
    Caja, CajaProducto, Inventario, Movimiento, MovimientoProducto,
    Carrito, CarritoItem, Pedido, Merma, AlertaStock, Notificacion,
    SugerenciaPedido, Conteo, ConteoProducto,
)
from .serializers import (
    SucursalSerializer, CategoriaSerializer, ProductoSerializer,
    ProveedorSerializer, ProductoProveedorSerializer,
    CajaSerializer, CajaProductoSerializer,
    InventarioSerializer,
    MovimientoSerializer, MovimientoCreateSerializer,
    CarritoSerializer, CarritoItemSerializer, PedidoSerializer,
    MermaSerializer, AlertaStockSerializer, NotificacionSerializer,
    SugerenciaPedidoSerializer, ConteoSerializer, ConteoProductoSerializer,
)


# ── CRUD estándar ─────────────────────────────────────────────────────────
class SucursalViewSet(viewsets.ModelViewSet):
    queryset = Sucursal.objects.all()
    serializer_class = SucursalSerializer
    filterset_fields = ["activa"]
    search_fields = ["nombre"]


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    search_fields = ["nombre"]


class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.select_related("categoria").all()
    serializer_class = ProductoSerializer
    filterset_fields = ["categoria", "activo", "sku"]
    search_fields = ["nombre"]


class ProveedorViewSet(viewsets.ModelViewSet):
    queryset = Proveedor.objects.all()
    serializer_class = ProveedorSerializer
    search_fields = ["nombre", "rut"]


class ProductoProveedorViewSet(viewsets.ModelViewSet):
    queryset = ProductoProveedor.objects.select_related("producto", "proveedor", "sucursal").all()
    serializer_class = ProductoProveedorSerializer
    filterset_fields = ["producto", "proveedor", "sucursal"]


class CajaViewSet(viewsets.ModelViewSet):
    queryset = Caja.objects.prefetch_related("contenido__producto").all()
    serializer_class = CajaSerializer
    filterset_fields = ["sku"]
    search_fields = ["nombre"]


class InventarioViewSet(viewsets.ModelViewSet):
    queryset = Inventario.objects.select_related("producto", "sucursal").all()
    serializer_class = InventarioSerializer
    filterset_fields = ["sucursal", "producto", "ubicacion"]
    search_fields = ["producto__nombre", "lote"]


class MovimientoViewSet(viewsets.ModelViewSet):
    queryset = Movimiento.objects.prefetch_related("productos__producto", "productos__caja_origen").all()
    serializer_class = MovimientoSerializer
    filterset_fields = ["sucursal", "tipo", "metodo", "tipo_sku"]
    ordering_fields = ["creado_en"]


class CarritoViewSet(viewsets.ModelViewSet):
    queryset = Carrito.objects.prefetch_related("items__producto").all()
    serializer_class = CarritoSerializer
    filterset_fields = ["sucursal"]


class CarritoItemViewSet(viewsets.ModelViewSet):
    queryset = CarritoItem.objects.select_related("producto").all()
    serializer_class = CarritoItemSerializer
    filterset_fields = ["carrito"]


class PedidoViewSet(viewsets.ModelViewSet):
    queryset = Pedido.objects.select_related("sucursal", "carrito").all()
    serializer_class = PedidoSerializer
    filterset_fields = ["sucursal", "estado"]
    ordering_fields = ["creado_en"]


class MermaViewSet(viewsets.ModelViewSet):
    queryset = Merma.objects.select_related("sucursal", "producto").all()
    serializer_class = MermaSerializer
    filterset_fields = ["sucursal", "razon"]
    ordering_fields = ["fecha"]


class AlertaStockViewSet(viewsets.ModelViewSet):
    queryset = AlertaStock.objects.select_related("sucursal", "producto").all()
    serializer_class = AlertaStockSerializer
    filterset_fields = ["sucursal", "tipo", "leida"]
    ordering_fields = ["creada_en"]


class NotificacionViewSet(viewsets.ModelViewSet):
    queryset = Notificacion.objects.select_related("sucursal").all()
    serializer_class = NotificacionSerializer
    filterset_fields = ["sucursal", "tipo", "leida"]
    ordering_fields = ["creada_en"]


class SugerenciaPedidoViewSet(viewsets.ModelViewSet):
    queryset = SugerenciaPedido.objects.select_related("producto", "proveedor", "sucursal").all()
    serializer_class = SugerenciaPedidoSerializer
    filterset_fields = ["sucursal", "estado"]
    ordering_fields = ["creada_en"]


class ConteoViewSet(viewsets.ModelViewSet):
    queryset = Conteo.objects.prefetch_related("productos__producto").all()
    serializer_class = ConteoSerializer
    filterset_fields = ["sucursal", "estado"]
    ordering_fields = ["creado_en"]

    @transaction.atomic
    @action(detail=True, methods=["post"])
    def confirmar(self, request, pk=None):
        conteo = self.get_object()
        if conteo.estado == "completado":
            return Response({"error": "Conteo ya completado"}, status=400)

        from datetime import datetime
        lote_base = f"AJ{datetime.now().strftime('%y%m%d%H%M')}"
        productos_ajustados = []

        for cp in conteo.productos.select_related("producto"):
            if cp.diferencia == 0:
                continue

            tipo = "entrada" if cp.diferencia > 0 else "salida"
            mov = Movimiento.objects.create(
                sucursal=conteo.sucursal,
                tipo=tipo,
                metodo="manual",
                sku_ingresado=cp.producto.sku,
                tipo_sku="producto",
                responsable=conteo.responsable,
                observaciones=f"Ajuste por conteo #{conteo.id}: sist={cp.cantidad_sistema} / fis={cp.cantidad_fisica}",
                control_calidad_ok=True,
            )
            MovimientoProducto.objects.create(
                movimiento=mov,
                producto=cp.producto,
                cantidad=abs(cp.diferencia),
            )

            inv, _ = Inventario.objects.get_or_create(
                sucursal=conteo.sucursal,
                producto=cp.producto,
                lote=lote_base,
                defaults={
                    "stock_actual": 0, "stock_minimo": Decimal("5"),
                    "fecha_recepcion": date.today(),
                },
            )
            inv.stock_actual += cp.diferencia
            if inv.stock_actual < 0:
                inv.stock_actual = Decimal("0")
            inv.save()

            cp.ajustado = True
            cp.save()
            productos_ajustados.append({
                "producto": cp.producto.nombre,
                "diferencia": float(cp.diferencia),
                "nuevo_stock": float(inv.stock_actual),
            })

        conteo.estado = "completado"
        conteo.completado_en = datetime.now()
        conteo.save()

        Notificacion.objects.create(
            sucursal=conteo.sucursal,
            tipo="sistema",
            titulo=f"Conteo #{conteo.id} completado",
            mensaje=f"Conteo #{conteo.id} en {conteo.sucursal.nombre}: {len(productos_ajustados)} productos ajustados",
            referencia_id=conteo.id,
        )

        return Response({"conteo_id": conteo.id, "productos_ajustados": productos_ajustados})


class ConteoProductoViewSet(mixins.CreateModelMixin, mixins.UpdateModelMixin, mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = ConteoProducto.objects.select_related("producto").all()
    serializer_class = ConteoProductoSerializer
    filterset_fields = ["conteo", "producto"]


# ── Endpoint: Ingreso/Salida por SKU ──────────────────────────────────────
@api_view(["POST"])
def movimiento_por_sku(request):
    """
    Crea un movimiento a partir del escaneo/ingreso de un SKU.
    El SKU puede ser de Producto o de Caja.
    Si es caja, se expande a sus productos constituyentes.
    """
    ser = MovimientoCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)

    data = ser.validated_data
    sucursal = Sucursal.objects.get(id=data["sucursal_id"])
    sku = data["sku"]
    tipo = data["tipo"]
    metodo = data["metodo"]
    responsable = data["responsable"]
    lote = data.get("lote", "") or f"L{datetime.now().strftime('%y%m%d%H%M')}"
    observaciones = data.get("observaciones", "")
    signo = 1 if tipo == "entrada" else -1

    # Control de calidad
    temperatura = data.get("temperatura")
    inspeccion_visual = data.get("inspeccion_visual", "")
    rechazar = data.get("rechazar", False)
    rechazo_motivo = data.get("rechazo_motivo", "")
    orden_compra_numero = data.get("orden_compra_numero", "")

    # Determinar si es producto o caja
    producto = Producto.objects.filter(sku=sku).first()
    caja = None
    tipo_sku = "producto"

    if not producto:
        caja = Caja.objects.filter(sku=sku).first()
        if caja:
            tipo_sku = "caja"
        else:
            return Response({"error": f"SKU '{sku}' no encontrado"}, status=404)

    # Validar temperatura segun tipo de conservacion del producto
    temp_ok = None
    if temperatura is not None and tipo == "entrada":
        temp = float(temperatura)
        if tipo_sku == "producto" and producto.tipo_conservacion:
            cons = producto.tipo_conservacion
            if cons == "refrigerado":
                temp_ok = 0 <= temp <= 5
                rango = "0-5°C"
            elif cons == "congelado":
                temp_ok = temp <= -18
                rango = "<= -18°C"
            else:
                temp_ok = True
                rango = "ambiente"
        else:
            # Si es caja o no tiene tipo, usamos regla general
            temp_ok = (0 <= temp <= 5) or (temp <= -18)
            rango = "Refrigerados: 0-5°C | Congelados: <= -18°C"

        if not temp_ok and not rechazar:
            return Response({
                "error": f"Temperatura {temperatura}°C fuera de rango para {producto.tipo_conservacion if tipo_sku=='producto' else 'este producto'}. Rango esperado: {rango}",
                "temp_actual": float(temperatura),
                "tipo_conservacion": producto.tipo_conservacion if tipo_sku == "producto" else "general",
                "rango_esperado": rango,
            }, status=400)

    control_calidad_ok = not rechazar if rechazar is not None else True

    # Si es rechazo, crear movimiento + notificacion a compras, no actualizar inventario
    if rechazar:
        mov = Movimiento.objects.create(
            sucursal=sucursal,
            tipo="entrada",
            metodo=metodo,
            sku_ingresado=sku,
            tipo_sku=tipo_sku,
            responsable=responsable,
            observaciones=observaciones,
            orden_compra_numero=orden_compra_numero,
            temperatura=temperatura,
            temp_ok=temp_ok,
            inspeccion_visual=inspeccion_visual,
            control_calidad_ok=False,
            rechazo_motivo=rechazo_motivo or "Rechazado en control de calidad",
        )
        # Notificar a compras
        Notificacion.objects.create(
            sucursal=sucursal,
            tipo="rechazo",
            titulo=f"Producto rechazado: {sku}",
            mensaje=(
                f"Movimiento #{mov.id}: {sku} rechazado por {rechazo_motivo or 'control de calidad'}. "
                f"Responsable: {responsable}. Temperatura: {temperatura}°C. "
                f"Inspeccion: {inspeccion_visual or 'N/A'}. OC: {orden_compra_numero or 'N/A'}"
            ),
            referencia_id=mov.id,
        )
        return Response(MovimientoSerializer(mov).data, status=201)

    with transaction.atomic():
        mov = Movimiento.objects.create(
            sucursal=sucursal,
            tipo=tipo,
            metodo=metodo,
            sku_ingresado=sku,
            tipo_sku=tipo_sku,
            responsable=responsable,
            observaciones=observaciones,
            temperatura=temperatura,
            temp_ok=temp_ok,
            inspeccion_visual=inspeccion_visual,
            orden_compra_numero=orden_compra_numero,
            control_calidad_ok=True,
        )

        if tipo_sku == "producto":
            cantidad = Decimal(str(request.data.get("cantidad", 1)))
            MovimientoProducto.objects.create(
                movimiento=mov,
                producto=producto,
                cantidad=cantidad,
            )
            inv, _ = Inventario.objects.get_or_create(
                sucursal=sucursal,
                producto=producto,
                lote=lote,
                defaults={
                    "stock_actual": 0,
                    "stock_minimo": Decimal("5"),
                    "fecha_recepcion": date.today(),
                },
            )
            inv.stock_actual += signo * cantidad
            inv.save()
            _generar_alerta(inv)

        else:  # caja
            for cp in caja.contenido.select_related("producto"):
                cantidad = cp.cantidad
                MovimientoProducto.objects.create(
                    movimiento=mov,
                    producto=cp.producto,
                    cantidad=cantidad,
                    caja_origen=caja,
                )
                inv, _ = Inventario.objects.get_or_create(
                    sucursal=sucursal,
                    producto=cp.producto,
                    lote=lote,
                    defaults={
                        "stock_actual": 0,
                        "stock_minimo": Decimal("5"),
                        "fecha_recepcion": date.today(),
                    },
                )
                inv.stock_actual += signo * cantidad
                inv.save()
                _generar_alerta(inv)

    return Response(MovimientoSerializer(mov).data, status=status.HTTP_201_CREATED)


@api_view(["POST"])
def movimiento_batch(request):
    """
    Crea un movimiento multiple (varios SKUs en una sola operacion).
    Para escaneo continuo: el usuario escanea N productos y los confirma todos juntos.
    Body: { sucursal_id, tipo, items: [{sku, cantidad, lote}], responsable, observaciones }
    """
    from django.db import transaction as tx

    sucursal_id = request.data.get("sucursal_id")
    tipo = request.data.get("tipo", "entrada")
    items = request.data.get("items", [])
    responsable = request.data.get("responsable", "")
    metodo = request.data.get("metodo", "manual")
    observaciones = request.data.get("observaciones", "")
    signo = 1 if tipo == "entrada" else -1

    if not items:
        return Response({"error": "Debe incluir al menos un item"}, status=400)

    try:
        sucursal = Sucursal.objects.get(id=sucursal_id)
    except Sucursal.DoesNotExist:
        return Response({"error": "Sucursal no encontrada"}, status=404)

    creados = []
    errores = []

    with tx.atomic():
        mov = Movimiento.objects.create(
            sucursal=sucursal,
            tipo=tipo,
            metodo=metodo,
            sku_ingresado=", ".join(i["sku"] for i in items[:5]) + ("..." if len(items) > 5 else ""),
            tipo_sku="producto",
            responsable=responsable,
            observaciones=observaciones,
            control_calidad_ok=True,
        )

        for item in items:
            sku = item.get("sku", "").strip()
            cantidad = Decimal(str(item.get("cantidad", 1)))
            lote = item.get("lote", "") or f"L{datetime.now().strftime('%y%m%d%H%M')}"

            producto = Producto.objects.filter(sku=sku).first()
            if not producto:
                # Intentar como caja
                caja = Caja.objects.filter(sku=sku).first()
                if caja:
                    for cp in caja.contenido.select_related("producto"):
                        MovimientoProducto.objects.create(
                            movimiento=mov,
                            producto=cp.producto,
                            cantidad=cp.cantidad,
                            caja_origen=caja,
                        )
                        inv, _ = Inventario.objects.get_or_create(
                            sucursal=sucursal, producto=cp.producto, lote=lote,
                            defaults={"stock_actual": 0, "stock_minimo": Decimal("5"), "fecha_recepcion": date.today()},
                        )
                        inv.stock_actual += signo * cp.cantidad
                        inv.save()
                        _generar_alerta(inv)
                    creados.append(f"Caja {sku}")
                    continue
                else:
                    errores.append(f"SKU '{sku}' no encontrado")
                    continue

            MovimientoProducto.objects.create(
                movimiento=mov, producto=producto, cantidad=cantidad,
            )
            inv, _ = Inventario.objects.get_or_create(
                sucursal=sucursal, producto=producto, lote=lote,
                defaults={"stock_actual": 0, "stock_minimo": Decimal("5"), "fecha_recepcion": date.today()},
            )
            inv.stock_actual += signo * cantidad
            inv.save()
            _generar_alerta(inv)
            creados.append(f"{producto.nombre} x {cantidad}")

    return Response({
        "movimiento": MovimientoSerializer(mov).data,
        "creados": creados,
        "errores": errores,
    }, status=201)


def _generar_alerta(inv):
    """Genera alerta si el stock bajó del minimo y sugiere reabastecimiento."""
    if inv.stock_actual <= 0:
        AlertaStock.objects.create(
            sucursal=inv.sucursal,
            producto=inv.producto,
            tipo="critico",
            stock_actual=inv.stock_actual,
            stock_minimo=inv.stock_minimo,
        )
        _generar_sugerencia(inv)
    elif inv.stock_actual < inv.stock_minimo:
        AlertaStock.objects.create(
            sucursal=inv.sucursal,
            producto=inv.producto,
            tipo="bajo",
            stock_actual=inv.stock_actual,
            stock_minimo=inv.stock_minimo,
        )
        _generar_sugerencia(inv)


def _generar_sugerencia(inv):
    """Crea una sugerencia de pedido si no hay una pendiente para ese producto/sucursal."""
    if SugerenciaPedido.objects.filter(
        sucursal=inv.sucursal, producto=inv.producto, estado="pendiente"
    ).exists():
        return
    proveedor = ProductoProveedor.objects.filter(
        producto=inv.producto, sucursal=inv.sucursal
    ).first()
    cantidad = inv.stock_minimo * Decimal("2") - inv.stock_actual  # sugerir el doble del deficit
    if cantidad > 0:
        SugerenciaPedido.objects.create(
            sucursal=inv.sucursal,
            producto=inv.producto,
            proveedor=proveedor.proveedor if proveedor else None,
            cantidad_sugerida=cantidad,
            stock_actual=inv.stock_actual,
            stock_minimo=inv.stock_minimo,
        )


# ── Dashboard Gerencia ────────────────────────────────────────────────────
@api_view(["GET"])
def dashboard_gerencia(request):
    sucursales = Sucursal.objects.filter(activa=True)
    total_suc = sucursales.count()
    total_prod = Producto.objects.filter(activo=True).count()
    total_stock = Inventario.objects.count()
    total_alertas = AlertaStock.objects.filter(leida=False).count()
    total_pedidos_pend = Pedido.objects.filter(estado="pendiente").count()

    # Mermas del mes
    inicio_mes = date.today().replace(day=1)
    total_mermas = Merma.objects.filter(fecha__gte=inicio_mes).count()

    # Valor total del inventario
    inventarios = Inventario.objects.select_related("producto").all()
    valor_total = sum(
        float(inv.stock_actual) * float(inv.producto.precio_venta)
        for inv in inventarios
    )

    # Ganancia potencial: precio_venta - precio_compra promedio
    ganancia = 0
    for inv in inventarios:
        pp = ProductoProveedor.objects.filter(
            producto=inv.producto, sucursal=inv.sucursal
        ).first()
        if pp:
            ganancia += float(inv.stock_actual) * (
                float(inv.producto.precio_venta) - float(pp.precio_compra)
            )

    # Stock por sucursal (top)
    stock_por_suc = []
    for s in sucursales:
        cant = Inventario.objects.filter(sucursal=s).aggregate(
            total=Sum("stock_actual")
        )["total"] or 0
        stock_por_suc.append({"id": s.id, "nombre": s.nombre, "total_stock": float(cant)})

    # Top 10 productos por valor en inventario
    top_prod = []
    prod_vals = {}
    for inv in inventarios:
        val = float(inv.stock_actual) * float(inv.producto.precio_venta)
        prod_vals[inv.producto.nombre] = prod_vals.get(inv.producto.nombre, 0) + val
    for nombre, val in sorted(prod_vals.items(), key=lambda x: -x[1])[:10]:
        top_prod.append({"nombre": nombre, "valor": round(val, 2)})

    return Response({
        "total_sucursales": total_suc,
        "total_productos": total_prod,
        "total_stock": total_stock,
        "total_alertas": total_alertas,
        "total_pedidos_pendientes": total_pedidos_pend,
        "total_mermas_mes": total_mermas,
        "valor_total_inventario": round(valor_total, 2),
        "ganancia_potencial": round(ganancia, 2),
        "stock_por_sucursal": stock_por_suc,
        "top_productos": top_prod,
    })


# ── Dashboard Admin (por sucursal) ────────────────────────────────────────
@api_view(["GET"])
def dashboard_admin(request, sucursal_id):
    suc = Sucursal.objects.get(id=sucursal_id)
    invs = Inventario.objects.filter(sucursal=suc).select_related("producto")

    total_prod = invs.count()
    total_stock = sum(float(i.stock_actual) for i in invs)
    total_alertas = AlertaStock.objects.filter(sucursal=suc, leida=False).count()
    total_pedidos = Pedido.objects.filter(sucursal=suc).count()
    total_movs = Movimiento.objects.filter(
        sucursal=suc, creado_en__gte=datetime.now() - timedelta(days=30)
    ).count()
    valor_inv = sum(float(i.stock_actual) * float(i.producto.precio_venta) for i in invs)

    # Stock bajo
    stock_bajo = []
    for inv in invs:
        if inv.stock_actual < inv.stock_minimo:
            stock_bajo.append({
                "producto": inv.producto.nombre,
                "sku": inv.producto.sku,
                "stock": float(inv.stock_actual),
                "minimo": float(inv.stock_minimo),
            })

    # Últimos movimientos
    ultimos = Movimiento.objects.filter(sucursal=suc).order_by("-creado_en")[:10]
    ultimos_data = [
        {
            "id": m.id,
            "tipo": m.tipo,
            "sku": m.sku_ingresado,
            "responsable": m.responsable,
            "fecha": m.creado_en.isoformat(),
        }
        for m in ultimos
    ]

    return Response({
        "sucursal_nombre": suc.nombre,
        "total_productos": total_prod,
        "total_stock": round(total_stock, 2),
        "total_alertas": total_alertas,
        "total_pedidos": total_pedidos,
        "total_movimientos_mes": total_movs,
        "valor_inventario": round(valor_inv, 2),
        "stock_bajo": stock_bajo,
        "ultimos_movimientos": ultimos_data,
    })


# ── Dashboard Bodeguero (por sucursal) ────────────────────────────────────
@api_view(["GET"])
def dashboard_bodeguero(request, sucursal_id):
    suc = Sucursal.objects.get(id=sucursal_id)
    invs = Inventario.objects.filter(sucursal=suc).select_related("producto")

    total_prod = invs.count()
    total_stock = sum(float(i.stock_actual) for i in invs)
    alertas_criticas = AlertaStock.objects.filter(
        sucursal=suc, leida=False, tipo="critico"
    ).count()

    # Stock por ubicación
    ubicaciones = {}
    for inv in invs:
        ubi = inv.get_ubicacion_display()
        ubicaciones[ubi] = ubicaciones.get(ubi, 0) + float(inv.stock_actual)
    stock_ubi = [{"ubicacion": k, "total": round(v, 2)} for k, v in ubicaciones.items()]

    # Movimientos de hoy
    hoy = date.today()
    movs_hoy = Movimiento.objects.filter(
        sucursal=suc, creado_en__date=hoy
    ).order_by("-creado_en")[:20]
    movs_data = [
        {
            "id": m.id,
            "tipo": m.tipo,
            "sku": m.sku_ingresado,
            "tipo_sku": m.tipo_sku,
            "metodo": m.metodo,
            "responsable": m.responsable,
            "hora": m.creado_en.strftime("%H:%M"),
        }
        for m in movs_hoy
    ]

    return Response({
        "sucursal_nombre": suc.nombre,
        "total_productos": total_prod,
        "total_stock": round(total_stock, 2),
        "alertas_criticas": alertas_criticas,
        "stock_por_ubicacion": stock_ubi,
        "movimientos_hoy": movs_data,
    })


# ── Aprobacion/Rechazo de pedidos ─────────────────────────────────────────
@api_view(["POST"])
def aprobar_pedido(request, pedido_id):
    """Aprueba un pedido y actualiza su estado a 'aprobado'."""
    try:
        pedido = Pedido.objects.get(id=pedido_id)
    except Pedido.DoesNotExist:
        return Response({"error": "Pedido no encontrado"}, status=404)

    if pedido.estado != "pendiente":
        return Response({"error": f"El pedido ya esta en estado '{pedido.estado}'"}, status=400)

    pedido.estado = "aprobado"
    pedido.aprobado_por = request.data.get("aprobado_por", request.data.get("responsable", ""))
    pedido.save()

    Notificacion.objects.create(
        sucursal=pedido.sucursal,
        tipo="pedido",
        titulo=f"Pedido #{pedido.id} aprobado",
        mensaje=f"Pedido #{pedido.id} de {pedido.sucursal.nombre} aprobado por {pedido.aprobado_por}",
        referencia_id=pedido.id,
    )

    return Response(PedidoSerializer(pedido).data)


@api_view(["POST"])
def rechazar_pedido(request, pedido_id):
    """Rechaza un pedido y lo marca como cancelado con motivo."""
    try:
        pedido = Pedido.objects.get(id=pedido_id)
    except Pedido.DoesNotExist:
        return Response({"error": "Pedido no encontrado"}, status=404)

    if pedido.estado != "pendiente":
        return Response({"error": f"El pedido ya esta en estado '{pedido.estado}'"}, status=400)

    motivo = request.data.get("motivo", "Sin motivo especificado")
    pedido.estado = "cancelado"
    pedido.rechazado_motivo = motivo
    pedido.save()

    Notificacion.objects.create(
        sucursal=pedido.sucursal,
        tipo="pedido",
        titulo=f"Pedido #{pedido.id} rechazado",
        mensaje=f"Pedido #{pedido.id} de {pedido.sucursal.nombre} rechazado: {motivo}",
        referencia_id=pedido.id,
    )

    return Response(PedidoSerializer(pedido).data)


@api_view(["POST"])
def avanzar_pedido(request, pedido_id):
    """Avanza un pedido al siguiente estado del flujo."""
    try:
        pedido = Pedido.objects.get(id=pedido_id)
    except Pedido.DoesNotExist:
        return Response({"error": "Pedido no encontrado"}, status=404)

    FLUJO = ["pendiente", "aprobado", "preparacion", "enviado", "recibido"]
    if pedido.estado not in FLUJO:
        return Response({"error": f"Estado '{pedido.estado}' no permite avance"}, status=400)

    idx = FLUJO.index(pedido.estado)
    if idx >= len(FLUJO) - 1:
        return Response({"error": "El pedido ya esta en el estado final"}, status=400)

    pedido.estado = FLUJO[idx + 1]
    pedido.save()

    Notificacion.objects.create(
        sucursal=pedido.sucursal,
        tipo="pedido",
        titulo=f"Pedido #{pedido.id} → {pedido.estado}",
        mensaje=f"Pedido #{pedido.id} de {pedido.sucursal.nombre} avanza a: {pedido.estado}",
        referencia_id=pedido.id,
    )

    return Response(PedidoSerializer(pedido).data)


# ── SSE: Notificaciones en tiempo real ────────────────────────────────
import queue as _queue

_sse_clients = []

def _notificar_sse(notificacion):
    import json
    data = json.dumps({
        "id": notificacion.id,
        "tipo": notificacion.tipo,
        "titulo": notificacion.titulo,
        "mensaje": notificacion.mensaje,
        "sucursal_id": notificacion.sucursal_id,
        "sucursal_nombre": notificacion.sucursal.nombre,
        "leida": False,
        "creada_en": notificacion.creada_en.isoformat(),
    })
    for q in _sse_clients[:]:
        try:
            q.put(data)
        except Exception:
            if q in _sse_clients:
                _sse_clients.remove(q)


from django.http import StreamingHttpResponse
from django.views.decorators.http import require_GET

@require_GET
def sse_stream(request):
    q = _queue.Queue()
    _sse_clients.append(q)

    def event_stream():
        try:
            yield "data: connected\n\n"
            while True:
                try:
                    data = q.get(timeout=30)
                    yield f"data: {data}\n\n"
                except _queue.Empty:
                    yield ": keepalive\n\n"
        except GeneratorExit:
            pass
        finally:
            if q in _sse_clients:
                _sse_clients.remove(q)

    return StreamingHttpResponse(
        event_stream(),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )


# Monkeypatch: cada Notificacion.create emite SSE
_create_notificacion = Notificacion.objects.create

def _notificacion_con_sse(**kwargs):
    n = _create_notificacion(**kwargs)
    _notificar_sse(n)
    return n

Notificacion.objects.create = _notificacion_con_sse


# ── Web Push: Suscripcion ────────────────────────────────────────────
from django.db import models as _models

class PushSubscription(_models.Model):
    endpoint = _models.URLField(max_length=500)
    auth = _models.CharField(max_length=100)
    p256dh = _models.CharField(max_length=200)
    user_agent = _models.CharField(max_length=300, blank=True)
    creada_en = _models.DateTimeField(auto_now_add=True)

    class Meta:
        app_label = "api"

    def __str__(self):
        return f"Push sub {self.id}"

@api_view(["POST"])
def push_subscribe(request):
    sub = PushSubscription.objects.create(
        endpoint=request.data.get("endpoint", ""),
        auth=request.data.get("keys", {}).get("auth", ""),
        p256dh=request.data.get("keys", {}).get("p256dh", ""),
        user_agent=request.META.get("HTTP_USER_AGENT", "")[:300],
    )
    return Response({"id": sub.id})


# ── Reportes: Mermas y Productos a Vencer ─────────────────────────────────
from django.http import HttpResponse
from datetime import date, timedelta

@api_view(["GET"])
def resumen_mermas(request):
    """Resumen de mermas: total por periodo, total por razon, perdida total."""
    desde = request.GET.get("desde", "")
    hasta = request.GET.get("hasta", "")
    sucursal_id = request.GET.get("sucursal_id")

    qs = Merma.objects.select_related("producto", "sucursal").all()
    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)

    total_perdida = sum(float(m.perdida) for m in qs)
    total_mermas = qs.count()

    # Agrupar por razon
    por_razon = {}
    for m in qs:
        por_razon[m.razon] = por_razon.get(m.razon, 0) + float(m.perdida)

    # Ultimas 20
    ultimas = [
        {
            "id": m.id, "sucursal": m.sucursal.nombre,
            "producto": m.producto.nombre, "sku": m.producto.sku,
            "cantidad": float(m.cantidad), "razon": m.razon,
            "responsable": m.responsable,
            "perdida": float(m.perdida),
            "fecha": m.fecha.isoformat(),
        }
        for m in qs.order_by("-fecha")[:20]
    ]

    return Response({
        "total_mermas": total_mermas,
        "total_perdida": round(total_perdida, 2),
        "por_razon": [{"razon": k, "perdida": round(v, 2)} for k, v in por_razon.items()],
        "ultimas": ultimas,
    })


@api_view(["GET"])
def productos_vencer(request):
    """Productos proximos a vencer (default: 7 dias)."""
    dias = int(request.GET.get("dias", 7))
    sucursal_id = request.GET.get("sucursal_id")
    hoy = date.today()
    limite = hoy + timedelta(days=dias)

    qs = Inventario.objects.select_related("producto", "sucursal").filter(
        fecha_caducidad__gte=hoy,
        fecha_caducidad__lte=limite,
        stock_actual__gt=0,
    )
    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)

    items = [
        {
            "id": i.id,
            "sucursal": i.sucursal.nombre,
            "producto": i.producto.nombre,
            "sku": i.producto.sku,
            "stock": float(i.stock_actual),
            "lote": i.lote,
            "caduca": i.fecha_caducidad.isoformat(),
            "dias_restantes": (i.fecha_caducidad - hoy).days,
            "perdida_potencial": round(float(i.stock_actual) * float(i.producto.precio_venta), 2),
        }
        for i in qs.order_by("fecha_caducidad")
    ]

    return Response({
        "dias": dias,
        "total_productos": len(items),
        "perdida_potencial": round(sum(i["perdida_potencial"] for i in items), 2),
        "items": items,
    })


# ── Descarga de Solicitud PDF/XLSX ───────────────────────────────────────

@api_view(["GET"])
def productos_con_proveedores(request):
    """Retorna productos agrupados por proveedor con datos de stock."""
    sucursal_id = request.GET.get("sucursal_id")
    solo_criticos = request.GET.get("criticos", "false").lower() == "true"

    pp_qs = ProductoProveedor.objects.select_related(
        "producto", "proveedor", "sucursal"
    ).all()

    if sucursal_id:
        pp_qs = pp_qs.filter(sucursal_id=sucursal_id)

    # Agrupar por proveedor
    proveedores_map = {}
    for pp in pp_qs:
        prov_id = pp.proveedor_id
        if prov_id not in proveedores_map:
            proveedores_map[prov_id] = {
                "id": prov_id,
                "nombre": pp.proveedor.nombre,
                "rut": pp.proveedor.rut,
                "contacto": pp.proveedor.contacto,
                "telefono": pp.proveedor.telefono,
                "productos": [],
            }

        # Obtener stock actual
        inv = Inventario.objects.filter(
            sucursal=pp.sucursal, producto=pp.producto
        ).first()
        stock = float(inv.stock_actual) if inv else 0
        minimo = float(inv.stock_minimo) if inv else 0

        if solo_criticos and stock >= minimo:
            continue

        proveedores_map[prov_id]["productos"].append({
            "id": pp.producto.id,
            "sku": pp.producto.sku,
            "nombre": pp.producto.nombre,
            "categoria": pp.producto.categoria.nombre if pp.producto.categoria else "",
            "unidad": pp.producto.unidad_medida,
            "precio_compra": float(pp.precio_compra),
            "stock_actual": stock,
            "stock_minimo": minimo,
            "critico": stock < minimo,
            "cantidad_sugerida": max(minimo * 2 - stock, 0),
        })

    result = [p for p in proveedores_map.values() if p["productos"]]
    return Response({"proveedores": result})


@api_view(["GET"])
def descargar_solicitud(request):
    """Descarga una solicitud de productos en PDF o XLSX."""
    formato = request.GET.get("formato", "pdf")
    sucursal_id = request.GET.get("sucursal_id")
    producto_ids = request.GET.get("productos", "")
    titulo = request.GET.get("titulo", "Solicitud de Productos")

    sucursal_nombre = "General"
    if sucursal_id:
        try:
            suc = Sucursal.objects.get(id=sucursal_id)
            sucursal_nombre = suc.nombre
        except Sucursal.DoesNotExist:
            pass

    ids = [int(x) for x in producto_ids.split(",") if x.strip().isdigit()]
    productos = Producto.objects.filter(id__in=ids).select_related("categoria")

    # Armar datos
    rows = []
    for prod in productos:
        pp = ProductoProveedor.objects.filter(
            producto=prod, sucursal_id=sucursal_id or 1
        ).first()
        inv = Inventario.objects.filter(
            producto=prod, sucursal_id=sucursal_id or 1
        ).first()
        rows.append({
            "sku": prod.sku,
            "producto": prod.nombre,
            "categoria": prod.categoria.nombre if prod.categoria else "",
            "unidad": prod.unidad_medida,
            "proveedor": pp.proveedor.nombre if pp else "—",
            "precio": float(pp.precio_compra) if pp else 0,
            "stock": float(inv.stock_actual) if inv else 0,
            "minimo": float(inv.stock_minimo) if inv else 0,
            "cantidad": max(
                (float(inv.stock_minimo) if inv else 0) * 2 - (float(inv.stock_actual) if inv else 0),
                0
            ),
        })

    if formato == "xlsx":
        return _generar_xlsx(titulo, sucursal_nombre, rows)
    return _generar_pdf(titulo, sucursal_nombre, rows)


def _generar_pdf(titulo, sucursal, rows):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )

    styles = getSampleStyleSheet()
    elements = []

    # Titulo
    elements.append(Paragraph(
        f"<b>{titulo}</b><br/>{sucursal}",
        styles["Title"]
    ))
    elements.append(Spacer(1, 0.3 * inch))

    # Tabla
    data = [[
        "SKU", "Producto", "Categoria", "Unidad", "Proveedor",
        "Precio", "Stock", "Minimo", "Solicitar"
    ]]
    for r in rows:
        data.append([
            r["sku"], r["producto"], r["categoria"], r["unidad"],
            r["proveedor"],
            f"$ {r['precio']:,.0f} CLP",
            str(r["stock"]), str(r["minimo"]), str(int(r["cantidad"]))
        ])

    if not rows:
        data.append(["—", "Sin productos seleccionados", "", "", "", "", "", "", ""])

    table = Table(data, colWidths=[65, 130, 65, 50, 100, 55, 45, 45, 55])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 0.3 * inch))
    from datetime import datetime
    elements.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total productos: {len(rows)}",
        styles["Normal"]
    ))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="solicitud_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


def _generar_xlsx(titulo, sucursal, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitud"

    # Fila 1: Titulo
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{titulo} — {sucursal}"
    ws["A1"].font = Font(bold=True, size=14, color="1a1a2e")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Fila 2: fecha
    from datetime import datetime
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="6b7280")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Encabezados
    headers = ["SKU", "Producto", "Categoria", "Unidad", "Proveedor",
               "Precio", "Stock", "Minimo", "Solicitar"]
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(bold=True, color="ffffff", size=10)
    thin_border = Border(
        left=Side(style="thin", color="e5e7eb"),
        right=Side(style="thin", color="e5e7eb"),
        top=Side(style="thin", color="e5e7eb"),
        bottom=Side(style="thin", color="e5e7eb"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Set column widths
    widths = [12, 30, 15, 8, 22, 10, 8, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = w

    # Datos
    light_fill = PatternFill(start_color="f9fafb", end_color="f9fafb", fill_type="solid")
    for i, r in enumerate(rows, 4):
        vals = [
            r["sku"], r["producto"], r["categoria"], r["unidad"],
            r["proveedor"],
            f"$ {r['precio']:,.0f}",
            r["stock"], r["minimo"], int(r["cantidad"])
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = light_fill

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="solicitud_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
